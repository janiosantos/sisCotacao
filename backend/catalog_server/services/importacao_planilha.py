"""Importação de lista de produtos por planilha CSV/XLSX.

Formato esperado: 1 linha de cabeçalho com as colunas (a principal pode ser
DESCRIÇÃO, DESCRICAO, NOME ou PRODUTO):

    DESCRICAO | MARCA | GRUPO | SUBGRUPO | CATEGORIA | SUBCATEGORIA | FAMILIA

- Colunas em qualquer ordem/caixa; acentos e espaços são normalizados.
- GRUPO/SUBGRUPO/CATEGORIA/SUBCATEGORIA/FAMILIA/MARCA são opcionais.
- A taxonomia é resolvida por nome (create-or-get): grupo, subgrupo, categoria,
  subcategoria, família e marca.
- Produtos são criados como RASCUNHO (ativo 0) — revisar/publicar antes de vender.
- Deduplicação por nome+marca (case-insensitive); linhas já existentes são
  contadas como "atualizadas" e não duplicam.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import re
import unicodedata
from pathlib import Path

from catalog_server import categorias
from catalog_server.db import system_conn
from catalog_server.repositories import marcas as marcas_repo

_CAMPOS = ("descricao", "marca", "grupo", "subgrupo", "categoria", "subcategoria", "familia")
_ALIASES = {"descricao": ("DESCRICAO", "NOME", "PRODUTO")}
_logger = logging.getLogger(__name__)


def _normalizar_nome(s) -> str:
    """Normaliza o rótulo da coluna: sem acentos/espaços, em maiúsculas."""
    s = unicodedata.normalize("NFD", (s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[\s_\-]+", "", s).upper()


def _valor(s):
    if s is None:
        return ""
    return str(s).strip()


def _aliases_campo(campo: str) -> tuple[str, ...]:
    return _ALIASES.get(campo, (_normalizar_nome(campo),))


def _ler_csv(conteudo: bytes) -> list[dict]:
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = conteudo.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Não foi possível decodificar o CSV (UTF-8/Latin-1).")

    primeira = (text.splitlines() or [""])[0]
    delim = ";" if primeira.count(";") >= primeira.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    if not reader.fieldnames:
        raise ValueError("Planilha sem cabeçalho.")
    mapa = {_normalizar_nome(h): h for h in reader.fieldnames}
    if not any(alias in mapa for alias in _ALIASES["descricao"]):
        raise ValueError("Cabeçalho sem a coluna obrigatória DESCRICAO, NOME ou PRODUTO.")

    def pega(raw: dict, campo: str) -> str:
        for alias in _aliases_campo(campo):
            valor = _valor(raw.get(mapa[alias])) if alias in mapa else ""
            if valor:
                return valor
        return ""

    linhas: list[dict] = []
    for raw in reader:
        linha = {campo: pega(raw, campo) for campo in _CAMPOS}
        if any(linha.values()):
            linhas.append(linha)
    return linhas


def _ler_xlsx(conteudo: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise ValueError("openpyxl não instalado — XLSX indisponível.")
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Arquivo XLSX sem planilhas.")
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise ValueError("Planilha XLSX vazia.")
    if not header:
        raise ValueError("Planilha XLSX sem cabeçalho.")
    mapa = {_normalizar_nome(h): i for i, h in enumerate(header) if h is not None}
    if not any(alias in mapa for alias in _ALIASES["descricao"]):
        raise ValueError("Cabeçalho sem a coluna obrigatória DESCRICAO, NOME ou PRODUTO.")
    linhas: list[dict] = []
    for r in rows:
        if r is None:
            continue
        def pega(campo: str) -> str:
            for alias in _aliases_campo(campo):
                idx = mapa.get(alias)
                if idx is not None and idx < len(r):
                    valor = _valor(r[idx])
                    if valor:
                        return valor
            return ""
        linha = {campo: pega(campo) for campo in _CAMPOS}
        if any(linha.values()):
            linhas.append(linha)
    return linhas


def ler_planilha(conteudo: bytes, nome_arquivo: str) -> list[dict]:
    """Lê a planilha (CSV ou XLSX) e devolve as linhas normalizadas."""
    ext = Path(nome_arquivo or "").suffix.lower()
    if ext == ".xlsx":
        return _ler_xlsx(conteudo)
    return _ler_csv(conteudo)


def _codigos_candidatos(nome: str, fallback: str):
    base = re.sub(r"[^A-Z0-9]", "", _normalizar_nome(nome))[:8] or fallback
    yield base
    for sequencia in range(2, 10_000):
        sufixo = str(sequencia)
        yield f"{base[: 8 - len(sufixo)]}{sufixo}"


def _resolver_grupo(conn, nome: str) -> int | None:
    if not nome:
        return None
    row = conn.execute(
        "SELECT id FROM grupos"
        " WHERE f_unaccent(LOWER(nome))=f_unaccent(LOWER(?))",
        (nome,),
    ).fetchone()
    if row:
        return row["id"]
    for codigo in _codigos_candidatos(nome, "G"):
        criado = conn.execute(
            "INSERT INTO grupos (codigo, nome, ativo) VALUES (?,?,1)"
            " ON CONFLICT DO NOTHING RETURNING id",
            (codigo, nome),
        ).fetchone()
        if criado:
            return criado["id"]
        # Outra importação pode ter criado o mesmo nome enquanto aguardávamos.
        row = conn.execute(
            "SELECT id FROM grupos"
            " WHERE f_unaccent(LOWER(nome))=f_unaccent(LOWER(?))",
            (nome,),
        ).fetchone()
        if row:
            return row["id"]
    raise ValueError(f"Não foi possível gerar um código único para o grupo '{nome}'.")


def _resolver_subgrupo(conn, grupo_id: int | None, nome: str) -> int | None:
    if not grupo_id or not nome:
        return None
    row = conn.execute(
        "SELECT id FROM subgrupos WHERE grupo_id=?"
        " AND f_unaccent(LOWER(nome))=f_unaccent(LOWER(?))",
        (grupo_id, nome),
    ).fetchone()
    if row:
        return row["id"]
    for codigo in _codigos_candidatos(nome, "S"):
        criado = conn.execute(
            "INSERT INTO subgrupos (grupo_id, codigo, nome, ativo) VALUES (?,?,?,1)"
            " ON CONFLICT DO NOTHING RETURNING id",
            (grupo_id, codigo, nome),
        ).fetchone()
        if criado:
            return criado["id"]
        row = conn.execute(
            "SELECT id FROM subgrupos"
            " WHERE grupo_id=?"
            " AND f_unaccent(LOWER(nome))=f_unaccent(LOWER(?))",
            (grupo_id, nome),
        ).fetchone()
        if row:
            return row["id"]
    raise ValueError(f"Não foi possível gerar um código único para o subgrupo '{nome}'.")


def _resolver_familia(conn, nome: str) -> int | None:
    if not nome:
        return None
    row = conn.execute(
        "SELECT id FROM familias"
        " WHERE f_unaccent(LOWER(nome))=f_unaccent(LOWER(?))",
        (nome,),
    ).fetchone()
    if row:
        return row["id"]
    return conn.execute(
        "INSERT INTO familias (nome, descricao) VALUES (?,?)",
        (nome, f"Importado de planilha ({nome})"),
    ).lastrowid


def _rejeicao(linha_numero: int, linha: dict, motivo: str, sugestao: str) -> dict:
    return {
        "linha": linha_numero,
        "status": "erro",
        "motivo": motivo,
        "sugestao": sugestao,
        "dados": {campo: _valor(linha.get(campo)) for campo in _CAMPOS},
    }


def _motivo_excecao(exc: Exception) -> tuple[str, str]:
    detalhe = str(exc).lower()
    if "unique constraint" in detalhe or "duplicate key" in detalhe:
        return (
            "Existe outro cadastro com a mesma identificação.",
            "Revise descrição, marca e classificação; remova duplicidades e importe novamente.",
        )
    if "foreign key" in detalhe:
        return (
            "A classificação informada não possui um cadastro pai válido.",
            "Confira GRUPO/SUBGRUPO e CATEGORIA/SUBCATEGORIA e importe novamente.",
        )
    if isinstance(exc, ValueError):
        return str(exc), "Corrija os valores indicados e importe esta planilha novamente."
    return (
        "Não foi possível cadastrar esta linha.",
        "Revise os campos da linha; se o problema continuar, encaminhe o relatório ao suporte.",
    )


def importar(
    conteudo: bytes,
    nome_arquivo: str,
    usuario_id: int | None = None,
) -> dict:
    """Importa a planilha: resolve taxonomia e cria produtos como rascunho."""
    linhas = ler_planilha(conteudo, nome_arquivo)
    if not linhas:
        raise ValueError("Planilha sem linhas de produtos para importar.")
    hash_conteudo = hashlib.sha256(conteudo).hexdigest()

    criados = 0
    atualizados = 0
    erros = 0
    erros_detalhe: list[dict] = []

    with system_conn() as conn:
        # Serializa o mesmo arquivo para impedir que dois envios simultâneos
        # criem produtos antes de o primeiro registrar a auditoria do lote.
        lock_key = int(hash_conteudo[:16], 16)
        if lock_key >= 2**63:
            lock_key -= 2**64
        conn.execute("SELECT pg_advisory_xact_lock(?)", (lock_key,))
        existente = conn.execute(
            "SELECT id, total, criados, atualizados, erros, resumo"
            " FROM cadastro_importacao WHERE hash_conteudo=?",
            (hash_conteudo,),
        ).fetchone()
        if existente:
            try:
                detalhes_existentes = json.loads(existente["resumo"] or "[]")
            except (TypeError, json.JSONDecodeError):
                detalhes_existentes = []
            return {
                "importacao_id": existente["id"],
                "duplicado": True,
                "total": existente["total"],
                "criados": existente["criados"],
                "atualizados": existente["atualizados"],
                "erros": existente["erros"],
                "erros_detalhe": detalhes_existentes,
                "relatorio_erros_url": (
                    f"/api/produtos-cadastro/importacoes/{existente['id']}/erros.xlsx"
                    if existente["erros"] else None
                ),
            }

        for idx, linha in enumerate(linhas, start=2):  # linha 1 = cabeçalho
            nome = (linha.get("descricao") or "").strip()
            marca = (linha.get("marca") or "").strip()
            if not nome:
                erros += 1
                erros_detalhe.append(_rejeicao(
                    idx,
                    linha,
                    "DESCRIÇÃO, DESCRICAO, NOME ou PRODUTO é obrigatório.",
                    "Preencha um dos campos aceitos e importe novamente.",
                ))
                continue

            existente = conn.execute(
                "SELECT id FROM produtos_cadastro"
                " WHERE f_unaccent(LOWER(nome))=f_unaccent(LOWER(?))"
                " AND f_unaccent(LOWER(COALESCE(marca,'')))=f_unaccent(LOWER(?))",
                (nome, marca),
            ).fetchone()
            if existente:
                atualizados += 1
                continue

            categoria = (linha.get("categoria") or "").strip()
            subcategoria = (linha.get("subcategoria") or "").strip()
            grupo = (linha.get("grupo") or "").strip()
            subgrupo = (linha.get("subgrupo") or "").strip()
            familia = (linha.get("familia") or "").strip()

            if subgrupo and not grupo:
                erros += 1
                erros_detalhe.append(_rejeicao(
                    idx,
                    linha,
                    "SUBGRUPO foi informado sem GRUPO.",
                    "Preencha o GRUPO correspondente ou remova o SUBGRUPO.",
                ))
                continue
            if subcategoria and not categoria:
                erros += 1
                erros_detalhe.append(_rejeicao(
                    idx,
                    linha,
                    "SUBCATEGORIA foi informada sem CATEGORIA.",
                    "Preencha a CATEGORIA correspondente ou remova a SUBCATEGORIA.",
                ))
                continue

            conn.execute("SAVEPOINT importar_produto_linha")
            try:
                categoria_id, subcategoria_id = categorias.resolve(conn, categoria, subcategoria)
                grupo_id = _resolver_grupo(conn, grupo)
                subgrupo_id = _resolver_subgrupo(conn, grupo_id, subgrupo)
                familia_id = _resolver_familia(conn, familia)
                marca_id = marcas_repo.resolver(conn, marca) if marca else None

                conn.execute(
                    "INSERT INTO produtos_cadastro "
                    "(familia_id, nome, marca, marca_id, categoria_id, subcategoria_id,"
                    " grupo_id, subgrupo_id, status_cadastro, ativo)"
                    " VALUES (?,?,?,?,?,?,?,?,?,0)",
                    (
                        familia_id,
                        nome,
                        marca or None,
                        marca_id,
                        categoria_id,
                        subcategoria_id,
                        grupo_id,
                        subgrupo_id,
                        "rascunho",
                    ),
                )
            except Exception as exc:  # noqa: BLE001 - rejeição isolada por linha
                conn.execute("ROLLBACK TO SAVEPOINT importar_produto_linha")
                conn.execute("RELEASE SAVEPOINT importar_produto_linha")
                motivo, sugestao = _motivo_excecao(exc)
                erros += 1
                erros_detalhe.append(_rejeicao(idx, linha, motivo, sugestao))
                _logger.warning("Linha %s rejeitada na importação de produtos: %s", idx, exc)
                continue
            conn.execute("RELEASE SAVEPOINT importar_produto_linha")
            criados += 1

        status = "erro" if erros and not (criados or atualizados) else "parcial" if erros else "ok"
        importacao = conn.execute(
            "INSERT INTO cadastro_importacao "
            "(arquivo_nome, hash_conteudo, total, criados, atualizados, erros, status, resumo, criado_por)"
            " VALUES (?,?,?,?,?,?,?,?,?)"
            " RETURNING id",
            (
                nome_arquivo,
                hash_conteudo,
                len(linhas),
                criados,
                atualizados,
                erros,
                status,
                json.dumps(erros_detalhe, ensure_ascii=False),
                usuario_id,
            ),
        ).fetchone()
    return {
        "importacao_id": importacao["id"],
        "duplicado": False,
        "total": len(linhas),
        "criados": criados,
        "atualizados": atualizados,
        "erros": erros,
        "erros_detalhe": erros_detalhe,
        "relatorio_erros_url": (
            f"/api/produtos-cadastro/importacoes/{importacao['id']}/erros.xlsx"
            if erros else None
        ),
    }


def gerar_planilha_erros(importacao_id: int) -> tuple[bytes, str] | None:
    """Gera uma planilha corrigível e reimportável com as linhas rejeitadas."""
    with system_conn() as conn:
        registro = conn.execute(
            "SELECT id, arquivo_nome, erros, resumo FROM cadastro_importacao WHERE id=?",
            (importacao_id,),
        ).fetchone()
    if not registro:
        return None

    try:
        rejeicoes = json.loads(registro["resumo"] or "[]")
    except (TypeError, json.JSONDecodeError):
        rejeicoes = []
    if not registro["erros"] or not rejeicoes:
        raise ValueError("Esta importação não possui linhas rejeitadas.")

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Nao importados"
    cabecalho = [campo.upper() for campo in _CAMPOS] + [
        "LINHA_ORIGINAL",
        "MOTIVO",
        "SUGESTAO",
    ]
    sheet.append(cabecalho)
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for rejeicao in rejeicoes:
        dados = rejeicao.get("dados") or {}
        valores = [_valor(dados.get(campo)) for campo in _CAMPOS]
        valores.extend([
            rejeicao.get("linha") or "",
            rejeicao.get("motivo") or "",
            rejeicao.get("sugestao") or "",
        ])
        sheet.append(valores)
        for cell in sheet[sheet.max_row]:
            if isinstance(cell.value, str):
                cell.data_type = "s"
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    larguras = [36, 20, 24, 24, 24, 24, 24, 14, 48, 58]
    for coluna, largura in enumerate(larguras, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(coluna)].width = largura

    instrucoes = workbook.create_sheet("Orientacoes")
    instrucoes.append(["Como corrigir e reimportar"])
    instrucoes["A1"].font = Font(bold=True, size=14, color="1F4E78")
    instrucoes.append(["1. Corrija os campos na aba 'Nao importados'."])
    instrucoes.append(["2. Use MOTIVO e SUGESTAO como orientação; essas colunas são ignoradas na importação."])
    instrucoes.append(["3. Salve o arquivo em XLSX e importe-o novamente no Cadastro de Produtos."])
    instrucoes.append(["4. Os produtos importados com sucesso no lote anterior não estão nesta planilha."])
    instrucoes.column_dimensions["A"].width = 110

    output = io.BytesIO()
    workbook.save(output)
    nome = f"produtos-nao-importados-{importacao_id}.xlsx"
    return output.getvalue(), nome
